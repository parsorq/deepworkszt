#!/usr/bin/env python3
"""
Builds an Excel apartment investment model with inputs, loan schedule,
cash flows, and summary metrics (IRR, ROI, MOIC, breakeven).
Run: pip install openpyxl && python build_model.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- CONFIG: Default assumptions (edit here or in the generated Excel) ---
DEFAULTS = {
    "Purchase price (local currency)": 3_500_000,
    "Down payment %": 0.30,
    "Closing costs % (of purchase)": 0.06,
    "Loan interest rate % (annual)": 10.0,
    "Loan term (years)": 20,
    "Monthly rent (year 1)": 25_000,
    "Rent growth % (annual)": 0.03,
    "Vacancy %": 0.05,
    "Operating expenses % (of gross rent)": 0.25,
    "Holding period (years)": 10,
    "Exit: sale price growth % (from purchase, per year)": 0.02,
    # Alternative: "Exit cap rate %" — if you prefer cap rate, model can use it
}

def style_header(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    return cell

def style_value(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        cell.number_format = "#,##0.00"
    return cell

def build_inputs_sheet(wb):
    ws = wb.active
    ws.title = "Inputs"
    style_header(ws, 1, 1, "Input")
    style_header(ws, 1, 2, "Value")
    style_header(ws, 1, 3, "Note")
    for i, (k, v) in enumerate(DEFAULTS.items(), start=2):
        ws.cell(row=i, column=1, value=k)
        style_value(ws, i, 2, v)
        ws.cell(row=i, column=3, value="")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 35
    return ws

def build_loan_sheet(wb, inputs_ws):
    ws = wb.create_sheet("Loan", 1)
    # Headers
    for c, h in enumerate(["Period", "Payment", "Principal", "Interest", "Balance"], 1):
        style_header(ws, 1, c, h)
    # Assumptions refs (same row as in Inputs)
    # Row 2 in Inputs: Purchase price, 3: Down %, 4: Closing %, 5: Rate, 6: Term
    price_ref = "Inputs!$B$2"
    down_pct_ref = "Inputs!$B$3"
    closing_pct_ref = "Inputs!$B$4"
    rate_ref = "Inputs!$B$5"
    term_ref = "Inputs!$B$6"
    loan_ref = f"({price_ref}*(1-{down_pct_ref}))"
    n_per_ref = f"{term_ref}*12"
    rate_month_ref = f"{rate_ref}/100/12"
    # Row 1: labels; Row 2: Period 1
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value=f"=-PMT({rate_month_ref},{n_per_ref},{loan_ref})")
    ws.cell(row=2, column=3, value=f"=PPMT({rate_month_ref},A2,{n_per_ref},{loan_ref})")
    ws.cell(row=2, column=4, value=f"=IPMT({rate_month_ref},A2,{n_per_ref},{loan_ref})")
    ws.cell(row=2, column=5, value=f"={loan_ref}+C2")
    for r in range(3, 22):  # placeholder rows; extend formula for full term
        period = r - 1
        ws.cell(row=r, column=1, value=period)
        ws.cell(row=r, column=2, value=f"=-PMT({rate_month_ref},{n_per_ref},{loan_ref})")
        ws.cell(row=r, column=3, value=f"=PPMT({rate_month_ref},A{r},{n_per_ref},{loan_ref})")
        ws.cell(row=r, column=4, value=f"=IPMT({rate_month_ref},A{r},{n_per_ref},{loan_ref})")
        ws.cell(row=r, column=5, value=f"=E{r-1}+C{r}")
    # Max period from term
    max_row = 2 + int(20 * 12)  # 20 years default; Excel will recalc from Inputs
    ws.column_dimensions["A"].width = 10
    for c in range(2, 6):
        ws.column_dimensions[get_column_letter(c)].width = 14
    return ws

def build_cashflow_sheet(wb, inputs_ws):
    ws = wb.create_sheet("Cash Flow", 2)
    # Headers: Year, Gross Rent, Vacancy, Effective Rent, OpEx, NOI, Debt Service, Levered CF, Sale Proceeds, Total CF, Cumulative CF
    headers = ["Year", "Gross Rent", "Vacancy", "Effective Rent", "OpEx", "NOI", "Debt Service", "Levered CF", "Sale Proceeds", "Total CF", "Cumulative CF"]
    for c, h in enumerate(headers, 1):
        style_header(ws, 1, c, h)
    price_ref = "Inputs!$B$2"
    down_ref = "Inputs!$B$3"
    close_ref = "Inputs!$B$4"
    rent_ref = "Inputs!$B$7"
    rent_grow_ref = "Inputs!$B$8"
    vac_ref = "Inputs!$B$9"
    opex_ref = "Inputs!$B$10"
    hold_ref = "Inputs!$B$11"
    exit_grow_ref = "Inputs!$B$12"
    rate_month = "Inputs!$B$5/100/12"
    term_months = "Inputs!$B$6*12"
    loan_amt = f"Inputs!$B$2*(1-Inputs!$B$3)"
    sale_price = f"{price_ref}*POWER(1+{exit_grow_ref},{hold_ref})"
    loan_bal_exit = f"FV({rate_month},{hold_ref}*12,PMT({rate_month},{term_months},-{loan_amt}),-{loan_amt})"
    equity_in = f"={price_ref}*{down_ref}+{price_ref}*{close_ref}"
    # Year 0: initial outflow
    ws.cell(row=2, column=1, value=0)
    ws.cell(row=2, column=2, value="Initial")
    ws.cell(row=2, column=8, value=f"=-({equity_in})")
    ws.cell(row=2, column=9, value=0)
    ws.cell(row=2, column=10, value=f"=H2+I2")
    ws.cell(row=2, column=11, value=f"=J2")
    # Year 1..Holding
    for r in range(3, 3 + 25):
        ws.cell(row=r, column=1, value=r - 2)
        ws.cell(row=r, column=2, value=f"=12*{rent_ref}*POWER(1+{rent_grow_ref},A{r}-1)")
        ws.cell(row=r, column=3, value=f"=-B{r}*{vac_ref}")
        ws.cell(row=r, column=4, value=f"=B{r}+C{r}")
        ws.cell(row=r, column=5, value=f"=-B{r}*{opex_ref}")
        ws.cell(row=r, column=6, value=f"=D{r}+E{r}")
        ws.cell(row=r, column=7, value="=-12*Loan!$B$2")
        ws.cell(row=r, column=8, value=f"=F{r}+G{r}")
        ws.cell(row=r, column=9, value=f"=IF(A{r}={hold_ref},{sale_price}-{loan_bal_exit},0)")
        ws.cell(row=r, column=10, value=f"=H{r}+I{r}")
        ws.cell(row=r, column=11, value=f"=K{r-1}+J{r}")
    ws.column_dimensions["A"].width = 8
    for c in range(2, 12):
        ws.column_dimensions[get_column_letter(c)].width = 14
    return ws

def build_exit_and_summary(wb, inputs_ws):
    ws = wb.create_sheet("Summary", 3)
    price_ref = "Inputs!$B$2"
    down_ref = "Inputs!$B$3"
    close_ref = "Inputs!$B$4"
    hold_ref = "Inputs!$B$11"
    exit_grow_ref = "Inputs!$B$12"
    equity_in = f"={price_ref}*{down_ref}+{price_ref}*{close_ref}"
    # Exit: sale price and loan payoff row
    ws.cell(row=1, column=1, value="Exit (end of holding period)")
    style_header(ws, 2, 1, "Item")
    style_header(ws, 2, 2, "Value")
    ws.cell(row=3, column=1, value="Sale price")
    ws.cell(row=3, column=2, value=f"={price_ref}*POWER(1+{exit_grow_ref},{hold_ref})")
    ws.cell(row=4, column=1, value="Loan balance at exit")
    # Loan balance at month = hold*12. Use CUMPRINC to get remaining balance or OFFSET into Loan sheet
    n_per = f"Inputs!$B$6*12"
    rate_month = f"Inputs!$B$5/100/12"
    loan_amt = f"Inputs!$B$2*(1-Inputs!$B$3)"
    ws.cell(row=4, column=2, value=f"=FV({rate_month},{n_per}-{hold_ref}*12,PMT({rate_month},{n_per},-{loan_amt}),-{loan_amt})")
    ws.cell(row=5, column=1, value="Net sale proceeds")
    ws.cell(row=5, column=2, value=f"=B3-B4")
    # Summary metrics
    ws.cell(row=7, column=1, value="Equity invested (down + closing)")
    ws.cell(row=7, column=2, value=f"={equity_in}")
    ws.cell(row=8, column=1, value="Total distributions (rental CF + sale)")
    ws.cell(row=8, column=2, value=f"=SUM(OFFSET('Cash Flow'!J2,1,0,{hold_ref},1))")
    ws.cell(row=9, column=1, value="Net sale proceeds (at exit)")
    ws.cell(row=9, column=2, value=f"=B5")
    ws.cell(row=10, column=1, value="Total distributions (same as B8)")
    ws.cell(row=10, column=2, value=f"=B8")
    ws.cell(row=11, column=1, value="MOIC (Multiple on Invested Capital)")
    ws.cell(row=11, column=2, value=f"=B10/B7")
    ws.cell(row=12, column=1, value="ROI %")
    ws.cell(row=12, column=2, value=f"=(B10-B7)/B7*100")
    ws.cell(row=13, column=1, value="IRR % (levered, annual)")
    ws.cell(row=13, column=2, value=f"=IRR(OFFSET('Cash Flow'!J2,0,0,{hold_ref}+1,1))*100")
    ws.cell(row=14, column=1, value="Breakeven year (first year cumulative CF >= 0)")
    ws.cell(row=14, column=2, value="=INDEX('Cash Flow'!A2:A25,MATCH(0,'Cash Flow'!K2:K25,1)+1)")
    for r in range(3, 15):
        ws.cell(row=r, column=2).number_format = "#,##0.00"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    return ws

def main():
    wb = openpyxl.Workbook()
    inputs_ws = build_inputs_sheet(wb)
    build_loan_sheet(wb, inputs_ws)
    build_cashflow_sheet(wb, inputs_ws)
    build_exit_and_summary(wb, inputs_ws)
    out_path = "apartment_investment_model.xlsx"
    wb.save(out_path)
    print(f"Saved {out_path}")

if __name__ == "__main__":
    main()
